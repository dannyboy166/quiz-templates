#!/usr/bin/env python3
"""
Generate the master question template spreadsheet for Kristie.

Creates QUESTION_TEMPLATE.xlsx with:
- Instructions sheet (how to fill in each column)
- Template sheet (21-column format with dropdowns)
- Reference Data sheet (valid values, ItemID numbering, examples)

Usage:
    python scripts/create_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = Path("data/questions/QUESTION_TEMPLATE.xlsx")

# Column definitions for the template sheet
COLUMNS = [
    # (header, width, description, validation_type)
    ("ItemID", 14, "Unique 8-digit question identifier", None),
    ("QuestionType", 14, "Select One, Select All, or True/False", "question_type"),
    ("Subject", 30, "Subject name (use dropdown)", "subject"),
    ("Category", 22, "SubjectArea / parent category", None),
    ("Grade", 8, "Year level (1-6)", "grade"),
    ("Topic", 22, "Topic name within the category", None),
    ("QuestionText", 50, "The question text shown to the student", None),
    ("Option1", 25, "Answer option 1 (leave blank for True/False)", None),
    ("Option2", 25, "Answer option 2 (leave blank for True/False)", None),
    ("Option3", 25, "Answer option 3", None),
    ("Option4", 25, "Answer option 4", None),
    ("Answer", 14, "Correct option: 1-4, 'True'/'False', or '1 and 3' for multiple", None),
    ("Level", 8, "Difficulty: 1=Easy, 2=Medium, 3=Hard", "level"),
    ("MediaType", 12, "P=Picture, VO=Voice Over, VONWQ=Voice Over No Written Question", "media"),
    ("ImageRequired", 14, "Y=question needs image to make sense, N=nice to have", "yesno"),
    ("ImageDescription", 20, "Image description/reference for Georgia", None),
    ("Hint1", 30, "First hint (easiest)", None),
    ("Hint2", 30, "Second hint (medium)", None),
    ("Hint3", 30, "Third hint (hardest/most helpful)", None),
    ("GetHelp", 40, "What teacher says/shows when student clicks Get Help", None),
    ("Notes", 30, "Internal notes (NOT imported to database)", None),
]

# Valid subject names (must match SUBJECT_MAP values in import_questions.py)
SUBJECTS = [
    "English",
    "Maths Number & Algebra",
    "Maths Measurement & Space",
    "Maths Statistics & Probability",
    "Human Society and its Environment",
    "Science & Technology",
    "Creative Arts",
    "Personal Development Health and Physical Education",
]

# Colours
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NOTES_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GETHELP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
INST_HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="2F5496")
INST_SUBHEADER_FONT = Font(name="Calibri", size=12, bold=True)
INST_BODY_FONT = Font(name="Calibri", size=11)
INST_CODE_FONT = Font(name="Consolas", size=11, color="C00000")
REF_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
REF_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
REF_ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Sample data rows
SAMPLE_ROWS = [
    # Select One - English
    {
        "ItemID": "20030001",
        "QuestionType": "Select One",
        "Subject": "English",
        "Category": "Grammar",
        "Grade": 3,
        "Topic": "Nouns",
        "QuestionText": "Which word is a noun?",
        "Option1": "quickly",
        "Option2": "beautiful",
        "Option3": "elephant",
        "Option4": "running",
        "Answer": 3,
        "Level": 1,
        "MediaType": "VO",
        "ImageRequired": "N",
        "ImageDescription": "",
        "Hint1": "A noun is a person, place, or thing",
        "Hint2": "Look for the word that names a thing",
        "Hint3": "An animal is a thing",
        "GetHelp": "The teacher says: A noun is a naming word. It can be a person like 'teacher', a place like 'school', or a thing like 'elephant'.",
        "Notes": "",
    },
    # Select All - Maths
    {
        "ItemID": "01030001",
        "QuestionType": "Select All",
        "Subject": "Maths Number & Algebra",
        "Category": "Addition",
        "Grade": 3,
        "Topic": "Addition",
        "QuestionText": "Which of these equal 10?",
        "Option1": "5 + 5",
        "Option2": "3 + 6",
        "Option3": "7 + 3",
        "Option4": "4 + 4",
        "Answer": "1 and 3",
        "Level": 2,
        "MediaType": "VO",
        "ImageRequired": "N",
        "ImageDescription": "",
        "Hint1": "Try adding each pair of numbers",
        "Hint2": "Two of these answers are correct",
        "Hint3": "5 + 5 = 10 and 7 + 3 = 10",
        "GetHelp": "The teacher says: To add two numbers, you can count on from the bigger number. For example, 7 + 3 means start at 7 and count 3 more: 8, 9, 10.",
        "Notes": "",
    },
    # True/False - Science
    {
        "ItemID": "00320001",
        "QuestionType": "True/False",
        "Subject": "Science & Technology",
        "Category": "Living Things",
        "Grade": 3,
        "Topic": "Classifying Living Things",
        "QuestionText": "A spider is an insect.",
        "Option1": "",
        "Option2": "",
        "Option3": "",
        "Option4": "",
        "Answer": "False",
        "Level": 2,
        "MediaType": "VO",
        "ImageRequired": "N",
        "ImageDescription": "",
        "Hint1": "Count the legs",
        "Hint2": "Insects have 6 legs",
        "Hint3": "Spiders have 8 legs - they are arachnids",
        "GetHelp": "The teacher says: Insects always have 6 legs and 3 body parts. Spiders have 8 legs, so they belong to a different group called arachnids.",
        "Notes": "",
    },
    # Select One - with image required
    {
        "ItemID": "01040001",
        "QuestionType": "Select One",
        "Subject": "Maths Measurement & Space",
        "Category": "3D",
        "Grade": 4,
        "Topic": "3D shapes",
        "QuestionText": "What is the name of this shape?",
        "Option1": "cube",
        "Option2": "sphere",
        "Option3": "cylinder",
        "Option4": "cone",
        "Answer": 3,
        "Level": 1,
        "MediaType": "VO",
        "ImageRequired": "Y",
        "ImageDescription": "Image of a cylinder",
        "Hint1": "Look at the shape carefully",
        "Hint2": "This shape has two flat circular faces",
        "Hint3": "It looks like a tin can",
        "GetHelp": "A video showing different 3D shapes spinning, with labels appearing on each one.",
        "Notes": "",
    },
    # VONWQ example
    {
        "ItemID": "20030002",
        "QuestionType": "Select One",
        "Subject": "English",
        "Category": "Phonics",
        "Grade": 3,
        "Topic": "Vowel sounds",
        "QuestionText": "Which word has the same vowel sound as 'cake'?",
        "Option1": "cat",
        "Option2": "rain",
        "Option3": "cup",
        "Option4": "dog",
        "Answer": 2,
        "Level": 2,
        "MediaType": "VONWQ",
        "ImageRequired": "N",
        "ImageDescription": "",
        "Hint1": "Say 'cake' out loud - what sound does the 'a' make?",
        "Hint2": "The 'a' in cake makes a long 'ay' sound",
        "Hint3": "'Rain' also has the long 'ay' sound",
        "GetHelp": "The teacher says: Some letters can make different sounds. The letter 'a' in 'cake' makes a long sound - 'ay'. Listen for the same sound in the other words.",
        "Notes": "VONWQ = student hears this question read aloud but doesn't see the text on screen",
    },
]


def create_instructions_sheet(ws):
    """Create the Instructions sheet with documentation."""
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "4472C4"
    ws.column_dimensions["A"].width = 90

    rows = [
        ("WORLDWISE QUESTION TEMPLATE - INSTRUCTIONS", INST_HEADER_FONT),
        ("", None),
        ("This template is used for creating questions for the WorldWise app.", INST_BODY_FONT),
        ("Use the 'Template' sheet to enter your questions. Use the 'Reference Data' sheet for valid values.", INST_BODY_FONT),
        ("", None),
        ("COLUMN GUIDE", INST_SUBHEADER_FONT),
        ("", None),
    ]

    # Column descriptions
    col_docs = [
        ("A - ItemID", "Unique 8-digit identifier for each question. See Reference Data sheet for numbering rules."),
        ("B - QuestionType", "Pick from the dropdown: 'Select One' (pick one answer), 'Select All' (pick all correct answers), or 'True/False'."),
        ("C - Subject", "The subject this question belongs to. Use the dropdown - must match exactly."),
        ("D - Category", "The category (SubjectArea) within the subject. E.g., 'Grammar' under English, 'Addition' under Maths."),
        ("E - Grade", "Year level: 1-6. This determines which Stage the question belongs to (Year 1-2 = Stage 1, Year 3-4 = Stage 2, Year 5-6 = Stage 3)."),
        ("F - Topic", "The specific topic within the category. E.g., 'Nouns' under Grammar, 'CVC Words' under Phonics."),
        ("G - QuestionText", "The question shown to the student. Write the full question here - do NOT embed answer options in the question text."),
        ("H-K - Option1 to Option4", "The answer choices. For True/False questions, leave these BLANK (the app shows True/False buttons automatically). For Select One/All, fill in at least 2 options."),
        ("L - Answer", "The correct answer. For Select One: put the option number (1, 2, 3, or 4). For Select All: put all correct numbers separated by 'and' (e.g., '1 and 3'). For True/False: put 'True' or 'False'."),
        ("M - Level", "Difficulty level: 1 = Easy, 2 = Medium, 3 = Hard."),
        ("N - MediaType", "P = Picture only (no voice over). VO = Voice Over (question is read aloud). VONWQ = Voice Over No Written Question (read aloud but text is hidden from screen)."),
        ("O - ImageRequired", "Y = the question NEEDS an image to make sense (e.g., 'What shape is this?'). N = an image would be nice but isn't essential."),
        ("P - ImageDescription", "Describe what image is needed for Georgia. This is a description, not a filename."),
        ("Q-S - Hint1 to Hint3", "Progressive hints shown when the student asks for help. Hint1 is the gentlest nudge, Hint3 gives the most help. Write each hint in its own column."),
        ("T - GetHelp", "What happens when the student clicks 'Get Help'. This could be text the teacher character says, OR a description of what should be shown (e.g., 'a video of the water cycle')."),
        ("U - Notes", "Internal notes for the team. These are NOT imported to the database."),
    ]

    for header, desc in col_docs:
        rows.append((header, INST_SUBHEADER_FONT))
        rows.append((desc, INST_BODY_FONT))
        rows.append(("", None))

    rows.extend([
        ("IMPORTANT RULES", INST_SUBHEADER_FONT),
        ("", None),
        ("1. Do NOT change the column order or add/remove columns.", INST_BODY_FONT),
        ("2. Do NOT embed answer options in the QuestionText (e.g., 'Is it a / b / c / d?'). Use the Option columns.", INST_BODY_FONT),
        ("3. Do NOT combine hints into one cell (e.g., 'H1 - hint H2 - hint'). Use separate Hint1, Hint2, Hint3 columns.", INST_BODY_FONT),
        ("4. For True/False questions, leave Option1-4 BLANK. The app shows True/False buttons automatically.", INST_BODY_FONT),
        ("5. For Select All questions, put all correct option numbers in the Answer column separated by 'and' (e.g., '1 and 3').", INST_BODY_FONT),
        ("6. Every question MUST have an ItemID, QuestionType, Subject, Category, Grade, QuestionText, and Answer.", INST_BODY_FONT),
        ("7. Use the DROPDOWNS for QuestionType, Subject, Grade, Level, MediaType, and ImageRequired. Do NOT type values manually.", INST_BODY_FONT),
        ("8. Every ItemID must be UNIQUE. Do not use the same ItemID for two different questions.", INST_BODY_FONT),
        ("9. Always fill in Category and MediaType — do not leave them blank.", INST_BODY_FONT),
        ("10. The Notes column is for your internal use only - it is NOT imported to the database.", INST_BODY_FONT),
        ("11. ImageDescription is a brief for Georgia describing what image is needed — NOT an image filename.", INST_BODY_FONT),
        ("", None),
        ("ITEMID NUMBERING", INST_SUBHEADER_FONT),
        ("", None),
        ("Format: [PP][GG][NNNN]  where PP = subject prefix, GG = year code, NNNN = sequence number", INST_CODE_FONT),
        ("", None),
        ("Maths:          01GGNNNN  (e.g., 01030001 = Maths Year 3, question 1)", INST_CODE_FONT),
        ("English:        02GGNNNN  (e.g., 02040001 = English Year 4, question 1)", INST_CODE_FONT),
        ("HSIE:           00G1NNNN  (e.g., 00310001 = HSIE Year 3, question 1)", INST_CODE_FONT),
        ("Sci & Tech:     00G2NNNN  (e.g., 00320001 = Science Year 3, question 1)", INST_CODE_FONT),
        ("Creative Arts:  00G3NNNN  (e.g., 00330001 = Creative Arts Year 3, question 1)", INST_CODE_FONT),
        ("PDHPE:          00G4NNNN  (e.g., 00340001 = PDHPE Year 3, question 1)", INST_CODE_FONT),
        ("", None),
        ("See the Reference Data sheet for the full numbering guide.", INST_BODY_FONT),
    ])

    for i, (text, font) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True)


def create_template_sheet(ws):
    """Create the Template sheet with headers, dropdowns, and sample data."""
    ws.title = "Template"
    ws.sheet_properties.tabColor = "70AD47"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Headers
    for col_idx, (name, width, desc, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validations
    validations = {
        "question_type": DataValidation(
            type="list", formula1='"Select One,Select All,True/False,Written,Sort,Link"', allow_blank=True,
            errorTitle="Invalid Question Type",
            error="Must be Select One, Select All, True/False, Written, Sort, or Link",
        ),
        "subject": DataValidation(
            type="list", formula1=f'"{",".join(SUBJECTS)}"', allow_blank=True,
            errorTitle="Invalid Subject",
            error="Must be one of the valid subject names. Check the Reference Data sheet.",
        ),
        "grade": DataValidation(
            type="list", formula1='"1,2,3,4,5,6"', allow_blank=True,
            errorTitle="Invalid Grade",
            error="Must be 1-6 (Year 1 to Year 6)",
        ),
        "level": DataValidation(
            type="list", formula1='"1,2,3"', allow_blank=True,
            errorTitle="Invalid Level",
            error="Must be 1 (Easy), 2 (Medium), or 3 (Hard)",
        ),
        "media": DataValidation(
            type="list", formula1='"P,VO,VONWQ"', allow_blank=True,
            errorTitle="Invalid MediaType",
            error="Must be P (Picture), VO (Voice Over), or VONWQ (Voice Over No Written Question)",
        ),
        "yesno": DataValidation(
            type="list", formula1='"Y,N"', allow_blank=True,
            errorTitle="Invalid value",
            error="Must be Y or N",
        ),
    }

    # Apply validations to columns (rows 2-5000)
    for col_idx, (_, _, _, val_type) in enumerate(COLUMNS, start=1):
        if val_type and val_type in validations:
            dv = validations[val_type]
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}5000")

    for dv in validations.values():
        ws.add_data_validation(dv)

    # Sample rows
    for row_idx, sample in enumerate(SAMPLE_ROWS, start=2):
        for col_idx, (col_name, _, _, _) in enumerate(COLUMNS, start=1):
            value = sample.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = SAMPLE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Special fill for GetHelp column
            if col_name == "GetHelp":
                cell.fill = GETHELP_FILL
            # Special fill for Notes column
            elif col_name == "Notes":
                cell.fill = NOTES_FILL

    # Format ItemID column as text (prevent Excel treating as number)
    for row in range(2, 5001):
        ws.cell(row=row, column=1).number_format = numbers.FORMAT_TEXT

    # Add description row as comments/note (row after samples)
    desc_row = len(SAMPLE_ROWS) + 3
    ws.cell(row=desc_row, column=1, value="--- Sample rows above. Delete them and start entering your questions below. ---").font = Font(
        name="Calibri", size=10, italic=True, color="808080"
    )


def create_reference_sheet(ws):
    """Create the Reference Data sheet."""
    ws.title = "Reference Data"
    ws.sheet_properties.tabColor = "ED7D31"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 50

    row = 1

    # --- Section 1: Valid Subjects ---
    def write_header(ws, row, headers):
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = REF_HEADER_FONT
            cell.fill = REF_HEADER_FILL
            cell.border = THIN_BORDER
        return row + 1

    def write_row(ws, row, values, alt=False):
        for col, text in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.border = THIN_BORDER
            if alt:
                cell.fill = REF_ALT_FILL
        return row + 1

    ws.cell(row=row, column=1, value="VALID SUBJECTS").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["Subject Name (use in column C)", "Example Categories", "Example Topics"])

    subject_examples = [
        ("English", "Grammar, Phonics, Punctuation, Reading Comprehension, Spelling, Vocabulary, Phonological Awareness", "Nouns, CVC Words, Full Stops, Inferential, Sight Words"),
        ("Maths Number & Algebra", "Addition, Subtraction, Multiplication, Division, Fractions, Decimals, Numbers, Money, Patterns", "Place value, Counting, Odd and Even Numbers"),
        ("Maths Measurement & Space", "Length, Mass, Volume, Time, Area, 2D, 3D, Angles, Position", "2D shapes, 3D shapes, Telling time"),
        ("Maths Statistics & Probability", "Data, Chance", "Reading graphs, Likelihood"),
        ("Human Society and its Environment", "Ancient Past, Australian Symbols, Geography, Celebrations", "Ancient Past - Greece, Australian Waterways"),
        ("Science & Technology", "Living Things, Forces, Light, Sound, Earth, Digital Systems", "Life Cycles, Push and Pull, The Sun"),
        ("Creative Arts", "Visual Arts, Music, Drama, Dance", "Primary colours, Rhythm patterns, Puppetry"),
        ("Personal Development Health and Physical Education", "Personal Safety, Healthy Lifestyles, Road Safety, Online Safety", "Sun Safety, Water Safety, Ball skills"),
    ]

    for i, (subj, cats, topics) in enumerate(subject_examples):
        row = write_row(ws, row, [subj, cats, topics], alt=(i % 2 == 1))

    row += 2

    # --- Section 2: Question Types ---
    ws.cell(row=row, column=1, value="QUESTION TYPES").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["QuestionType (dropdown)", "What it means", "How to fill in"])

    template_info = [
        ("Select One", "Student picks ONE correct answer", "Fill in Option1-4 (at least 2). Answer = one number (e.g., 3)."),
        ("Select All", "Student picks ALL correct answers", "Fill in Option1-4 (at least 2). Answer = numbers with 'and' (e.g., '1 and 3')."),
        ("True/False", "Student picks True or False", "Leave Option1-4 BLANK. Answer = 'True' or 'False'."),
        ("Written", "Student types a written answer", "Leave Option1-4 BLANK. Answer = the correct text."),
        ("Sort", "Student drags items into the correct order", "Fill in Option1-4 with items to sort. Answer = correct order (e.g., '3,1,4,2')."),
        ("Link", "Student matches/links items together", "Fill in Option1-4 with items. Answer = matching pairs."),
    ]

    for i, (qtype, meaning, howto) in enumerate(template_info):
        row = write_row(ws, row, [qtype, meaning, howto], alt=(i % 2 == 1))

    row += 2

    # --- Section 3: Media Types ---
    ws.cell(row=row, column=1, value="MEDIA TYPES").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["Code", "Name", "What it means"])

    media_info = [
        ("P", "Picture", "Question has a picture but no voice over."),
        ("VO", "Voice Over", "Question is read aloud to the student AND shown on screen."),
        ("VONWQ", "Voice Over No Written Question", "Question is read aloud but the text is HIDDEN from screen. Student only hears it."),
    ]

    for i, (code, name, desc) in enumerate(media_info):
        row = write_row(ws, row, [code, name, desc], alt=(i % 2 == 1))

    row += 2

    # --- Section 4: Difficulty Levels ---
    ws.cell(row=row, column=1, value="DIFFICULTY LEVELS").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["Level", "Name", "Description"])

    level_info = [
        ("1", "Easy", "Basic recall or simple application of a concept."),
        ("2", "Medium", "Requires some thinking or application."),
        ("3", "Hard", "Requires deeper understanding or multi-step reasoning."),
    ]

    for i, (lev, name, desc) in enumerate(level_info):
        row = write_row(ws, row, [lev, name, desc], alt=(i % 2 == 1))

    row += 2

    # --- Section 5: Grades and Stages ---
    ws.cell(row=row, column=1, value="GRADES AND STAGES").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["Grade (Year)", "Stage", "Notes"])

    grade_info = [
        ("1", "Stage 1", "Year 1"),
        ("2", "Stage 1", "Year 2"),
        ("3", "Stage 2", "Year 3"),
        ("4", "Stage 2", "Year 4"),
        ("5", "Stage 3", "Year 5"),
        ("6", "Stage 3", "Year 6"),
    ]

    for i, (grade, stage, notes) in enumerate(grade_info):
        row = write_row(ws, row, [grade, stage, notes], alt=(i % 2 == 1))

    row += 2

    # --- Section 6: ItemID Numbering ---
    ws.cell(row=row, column=1, value="ITEMID NUMBERING GUIDE").font = INST_SUBHEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Format: [PREFIX][YEAR][SEQUENCE]").font = INST_CODE_FONT
    row += 2

    row = write_header(ws, row, ["Subject", "Prefix", "Year Code", "Example"])

    id_info = [
        ("Maths", "01", "GG (01-06)", "01030001 = Maths Year 3, question 1"),
        ("English", "02", "GG (01-06)", "02040001 = English Year 4, question 1"),
        ("HSIE", "00G1", "G = grade digit", "00310001 = HSIE Year 3, question 1"),
        ("Science & Technology", "00G2", "G = grade digit", "00320001 = Science Year 3, question 1"),
        ("Creative Arts", "00G3", "G = grade digit", "00330001 = Creative Arts Year 3, question 1"),
        ("PDHPE", "00G4", "G = grade digit", "00340001 = PDHPE Year 3, question 1"),
    ]

    for i, (subj, prefix, year_code, example) in enumerate(id_info):
        row = write_row(ws, row, [subj, prefix, year_code, example], alt=(i % 2 == 1))

    row += 2

    # --- Section 7: ImageRequired ---
    ws.cell(row=row, column=1, value="IMAGE REQUIRED GUIDE").font = INST_SUBHEADER_FONT
    row += 1
    row = write_header(ws, row, ["Value", "Meaning", "Example"])

    image_info = [
        ("Y", "Question NEEDS the image to make sense. Priority for image creation.", "'What shape is this?' - student can't answer without seeing the shape"),
        ("N", "Image is a nice-to-have enhancement. Question works without it.", "'What is 2 + 3?' - student can answer without an image"),
    ]

    for i, (val, meaning, example) in enumerate(image_info):
        row = write_row(ws, row, [val, meaning, example], alt=(i % 2 == 1))

    row += 2

    # --- Section 8: GetHelp Guide ---
    ws.cell(row=row, column=1, value="GET HELP COLUMN GUIDE").font = INST_SUBHEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="The GetHelp column describes what happens when a student clicks the 'Get Help' button.").font = INST_BODY_FONT
    row += 1
    ws.cell(row=row, column=1, value="This is different from Hints - hints are small nudges, Get Help is a full explanation.").font = INST_BODY_FONT
    row += 2

    row = write_header(ws, row, ["Type", "How to write it", "Example"])

    gethelp_info = [
        ("Teacher speaks", "Start with: The teacher says:", "The teacher says: A noun is a naming word for a person, place, or thing."),
        ("Visual/video needed", "Describe what should be shown", "A video showing the water cycle - evaporation, condensation, precipitation."),
        ("Interactive demo", "Describe the concept to demonstrate", "An animation showing how fractions work by dividing a pizza into equal parts."),
    ]

    for i, (gtype, howto, example) in enumerate(gethelp_info):
        row = write_row(ws, row, [gtype, howto, example], alt=(i % 2 == 1))


def main():
    wb = Workbook()

    # Sheet 1: Instructions
    ws_instructions = wb.active
    create_instructions_sheet(ws_instructions)

    # Sheet 2: Template
    ws_template = wb.create_sheet()
    create_template_sheet(ws_template)

    # Sheet 3: Reference Data
    ws_reference = wb.create_sheet()
    create_reference_sheet(ws_reference)

    # Set Template as the active sheet (what opens first)
    wb.active = 1

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template saved to: {OUTPUT_PATH}")
    print(f"  - Instructions sheet: column guide, rules, ItemID numbering")
    print(f"  - Template sheet: 21 columns with dropdowns and {len(SAMPLE_ROWS)} sample rows")
    print(f"  - Reference Data sheet: valid values for all dropdown fields")


if __name__ == "__main__":
    main()
