"""
Apply spelling/typo fixes to Stage 1 clean spreadsheets.

Reads from review_categorised.json (77 major issues) and review_filtered_report.json (1006 minor issues).
Skips:
- Wrong answer items (need Julie's approval)
- Intentional misspellings (spelling/phonics questions with deliberate wrong options)
- Australian English that's correct (spelt, colour, recognise etc.)

Usage:
    source venv/bin/activate
    python scripts/fix_spreadsheets.py --dry-run   # Preview changes
    python scripts/fix_spreadsheets.py              # Apply changes
"""

import json
import re
import sys
import os
from collections import defaultdict
from copy import copy

import openpyxl

# === Configuration ===

DATA_DIR = "data/questions"
CLEAN_DIR = os.path.join(DATA_DIR, "clean")

# Map sheet names to files
SHEET_TO_FILE = {
    "Phonological Awareness": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Phonics": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Spelling": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Punctuation": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Grammar": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Vocabulary": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Reading Comprehension": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Number and Algebra": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "Measurement & Space": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "Statistics & Probability": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "HSIE": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "Science & technology": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "Creative Arts": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "PD H PE": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
}

# Column name to index (1-based)
COL_MAP = {
    "ItemID": 1,
    "QuestionType": 2,
    "Subject": 3,
    "Category": 4,
    "Grade": 5,
    "Topic": 6,
    "QuestionText": 7,
    "Option1": 8,
    "Option2": 9,
    "Option3": 10,
    "Option4": 11,
    "Answer": 12,
    "Level": 13,
    "MediaType": 14,
    "ImageRequired": 15,
    "ImageDescription": 16,
    "Hint1": 17,
    "Hint2": 18,
    "Hint3": 19,
    "GetHelp": 20,
    "Notes": 21,
}

# Items with wrong answers - need Julie, skip entirely
WRONG_ANSWER_IDS = {"20023618", "20231810", "20012001"}

# Specific spelling entries to skip (intentional misspellings)
SKIP_ITEM_SPELLINGS = {
    # Apostrophe questions - missing apostrophe IS the question
    ("20209039", "youre", "you're"),
    ("20209043", "Id", "I'd"),
    ("20209046", "shouldnt", "shouldn't"),
    # Phonics 'which word says' - wrong spellings are distractors
    ("20231810", "disapere", "disappear"),
    ("20231810", "disappere", "disappear"),
    ("20231810", "disappeer", "disappear"),
    # Capital letter question - lowercase is intentional
    ("20205705", "mr smith", "Mr Smith"),
    # Homophone question - be/bee distinction is the point
    ("20205967", "be", "bee"),
    # 's' -> 'Its' is too ambiguous - handled as special case below
    ("00220630", "s", "Its"),
    # 'leaf' vs 'leaves' - debatable, leave for Julie
    ("00221421", "leaf", "leaves"),
    # These produce bad replacements - handle manually below
    ("00240238", "jumpa", "jump around"),  # "jumpa round" needs special handling
    ("00140819", "jumpa", "jump a"),  # "jumpa round" needs special handling
    ("00141229", "nights", "night"),  # Should be "night's" not "night"
    ("00141618", "word", "words"),  # Hits "password" instead of target
    ("00240422", "suns harmul UV rays", "sun"),  # Truncates the sentence
}


def is_intentional_misspelling(item_id, sheet, question_text, col, old_text, new_text):
    """Check if a proposed fix is actually an intentional misspelling."""
    q = question_text.lower() if question_text else ""

    # Check explicit skip list
    if (item_id, old_text, new_text) in SKIP_ITEM_SPELLINGS:
        return True

    # Spelling/phonics questions - options contain intentional wrong spellings
    is_option_col = col in ("Option1", "Option2", "Option3", "Option4")
    if is_option_col:
        intentional_patterns = [
            "spelt incorrectly",
            "correct spelling",
            "which word says",
            "needs an apostrophe",
            "needs a capital",
            "needing a capital",
            "which two words rhyme",
            "not a real word",
            "not a word",
            "means more than one",
            "choose the word that",
            "which is the correct",
        ]
        if any(p in q for p in intentional_patterns):
            return True

    # Questions asking to find incorrect words - the wrong words in the
    # question text itself are intentional
    is_question_col = col == "QuestionText"
    if is_question_col:
        question_with_deliberate_errors = [
            "written incorrectly",
            "written incorre",  # truncated in JSON report
            "spelt incorrectly",
            "spelt incorre",
            "correct spelling",
        ]
        if any(p in q for p in question_with_deliberate_errors):
            return True

    return False


def build_fixes():
    """Build the complete list of fixes from both review files."""
    with open(os.path.join(DATA_DIR, "review_categorised.json")) as f:
        categorised = json.load(f)

    with open(os.path.join(DATA_DIR, "review_filtered_report.json")) as f:
        filtered = json.load(f)

    fixes = []  # (item_id, sheet, col, old_text, new_text, severity)

    # --- Manual fixes for cases the AI report got wrong/ambiguous ---
    # "It's position" -> "Its position" (possessive, not contraction)
    fixes.append(("00220630", "Science & technology", "Option3",
                  "It's position", "Its position", "major"))
    # "oclock" -> "o'clock" - regex can't parse apostrophe in replacement
    for col in ["QuestionText", "Option1", "Option3", "Option4"]:
        fixes.append(("10012120", "Measurement & Space", col,
                      "oclock", "o'clock", "major"))
    # Missing 'to' - "go the park" -> "go to the park"
    fixes.append(("20041407", "Grammar", "QuestionText",
                  "go the park", "go to the park", "major"))
    # Pronoun mismatch - "helps  her" -> "helps  him" (Tom is male)
    fixes.append(("00121206", "Science & technology", "QuestionText",
                  "helps  her", "helps  him", "major"))
    # "jumpa round" -> "jump around" (both items have same text)
    fixes.append(("00240238", "PD H PE", "Option1",
                  "jumpa round.", "jump around.", "major"))
    fixes.append(("00140819", "PD H PE", "Option1",
                  "jumpa round.", "jump around.", "major"))
    # "nights sleep" -> "night's sleep" (possessive)
    fixes.append(("00141229", "PD H PE", "QuestionText",
                  "good nights sleep", "good night's sleep", "major"))
    # "a few word" -> "a few words" (not "password")
    fixes.append(("00141618", "PD H PE", "Option2",
                  "a few word ", "a few words ", "major"))
    # "suns harmul" -> "sun's harmful" (possessive + spelling)
    fixes.append(("00240422", "PD H PE", "QuestionText",
                  "suns harmul", "sun's harmful", "major"))

    # --- From real_fixes (major) ---
    for item in categorised["real_fixes"]:
        item_id = item["item_id"]
        sheet = item["sheet"]
        question = item.get("question", "")

        if item_id in WRONG_ANSWER_IDS:
            continue

        for sp in item.get("spelling", []):
            col, old_text, new_text = parse_spelling_entry(sp)
            if not old_text or not new_text or old_text == new_text:
                continue

            if is_intentional_misspelling(item_id, sheet, question, col, old_text, new_text):
                continue

            fixes.append((item_id, sheet, col, old_text, new_text, "major"))

    # --- From minor_issues ---
    for item in filtered["minor_issues"]:
        item_id = item["item_id"]
        sheet = item["sheet"]
        question = item.get("question", "")

        for sp in item.get("spelling", []):
            col, old_text, new_text = parse_spelling_entry(sp)
            if not old_text or not new_text or old_text == new_text:
                continue

            if is_intentional_misspelling(item_id, sheet, question, col, old_text, new_text):
                continue

            fixes.append((item_id, sheet, col, old_text, new_text, "minor"))

    return fixes


def parse_spelling_entry(sp):
    """Parse a spelling entry string into (column, old_text, new_text)."""
    sp_lower = sp.lower()

    # Determine column
    if "hint 1" in sp_lower:
        col = "Hint1"
    elif "hint 2" in sp_lower:
        col = "Hint2"
    elif "hint 3" in sp_lower:
        col = "Hint3"
    elif "option 1" in sp_lower:
        col = "Option1"
    elif "option 2" in sp_lower:
        col = "Option2"
    elif "option 3" in sp_lower:
        col = "Option3"
    elif "option 4" in sp_lower:
        col = "Option4"
    elif "question" in sp_lower:
        col = "QuestionText"
    else:
        # Try to infer from content
        col = "unknown"

    # Extract old -> new
    # Try multiple regex patterns for different AI report formats:
    patterns = [
        # 'X' should be 'Y' / 'X' should be spelled 'Y'
        r"'([^']+)'\s+should be\s+(?:spelled\s+|spelt\s+)?'([^']+)'",
        # X should be 'Y' (no quotes on old, old must be a word-like string)
        r"\b([A-Za-z]{2,}[A-Za-z ]*[A-Za-z])\s+should be\s+'([^']+)'",
        # X should be Y (no quotes at all, both must be word-like, max 3 words each)
        r"\b([A-Za-z]{2,}(?:\s[A-Za-z]+){0,2})\s+should be\s+([A-Za-z]{2,}(?:\s[A-Za-z]+){0,2})(?:\s|$|[,.(])",
    ]
    for pattern in patterns:
        m = re.search(pattern, sp)
        if m:
            old_text = m.group(1)
            new_text = m.group(2)
            # Sanity check: if new_text is suspiciously short compared to old_text,
            # it might be a regex parsing issue (e.g. "o'clock" parsed as "o")
            if len(new_text) < len(old_text) // 2 and len(new_text) <= 2:
                continue  # Try next pattern
            # Skip if old == new
            if old_text == new_text:
                continue
            # Skip if old_text contains the word "should" (regex matched wrong part)
            if "should" in old_text.lower():
                continue
            # Skip if new_text contains descriptor words (regex captured description)
            garbage_words = ["spelled", "capitalised", "capitalis", "in option",
                           "in hint", "in question", "in the", "original"]
            if any(g in new_text.lower() for g in garbage_words):
                continue
            # Skip if new_text is much longer than old_text (likely captured context)
            if len(new_text) > len(old_text) * 2 + 5:
                continue
            return col, old_text, new_text

    return col, None, None


def find_row_by_itemid(ws, item_id):
    """Find the row number for a given ItemID in a worksheet."""
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell_val = str(row[0].value).strip() if row[0].value else ""
        if cell_val == item_id:
            return row[0].row
    return None


def safe_replace(cell_value, old_text, new_text):
    """
    Replace old_text with new_text in cell_value, using word-boundary matching
    for short strings to avoid false positives (e.g. 'ot' in 'not').
    """
    # For short strings (<=3 chars), use word-boundary regex to avoid partial matches
    if len(old_text) <= 3:
        escaped = re.escape(old_text)
        pattern = r'(?<!\w)' + escaped + r'(?!\w)'
        match = re.search(pattern, cell_value)
        if match:
            return cell_value[:match.start()] + new_text + cell_value[match.end():]
        return None
    else:
        if old_text in cell_value:
            return cell_value.replace(old_text, new_text, 1)
        return None


def apply_fix(ws, row, col_name, old_text, new_text):
    """Apply a text replacement in a cell. Returns (success, detail)."""
    col_idx = COL_MAP.get(col_name)
    if not col_idx:
        return False, f"Unknown column: {col_name}"

    cell = ws.cell(row=row, column=col_idx)
    cell_value = str(cell.value) if cell.value else ""

    result = safe_replace(cell_value, old_text, new_text)
    if result is not None:
        cell.value = result
        return True, f"Fixed '{old_text}' -> '{new_text}'"
    else:
        return False, f"'{old_text}' not found in cell (value: '{cell_value[:50]}')"


def main():
    dry_run = "--dry-run" in sys.argv

    print("=" * 60)
    print("SPREADSHEET FIX SCRIPT")
    print(f"Mode: {'DRY RUN (no changes)' if dry_run else 'APPLYING FIXES'}")
    print("=" * 60)
    print()

    # Build fixes
    fixes = build_fixes()
    print(f"Total fixes to apply: {len(fixes)}")

    # Group by file
    by_file = defaultdict(list)
    for fix in fixes:
        item_id, sheet, col, old_text, new_text, severity = fix
        filename = SHEET_TO_FILE.get(sheet)
        if filename:
            by_file[filename].append(fix)
        else:
            print(f"  WARNING: Unknown sheet '{sheet}' for item {item_id}")

    for filename, file_fixes in sorted(by_file.items()):
        print(f"\n  {filename}: {len(file_fixes)} fixes")

    print()

    # Process each file
    total_applied = 0
    total_skipped = 0
    total_not_found = 0
    results = []

    for filename, file_fixes in sorted(by_file.items()):
        filepath = os.path.join(CLEAN_DIR, filename)
        print(f"\n{'='*60}")
        print(f"Processing: {filename}")
        print(f"{'='*60}")

        wb = openpyxl.load_workbook(filepath)

        # Group fixes by sheet
        sheet_fixes = defaultdict(list)
        for fix in file_fixes:
            sheet_fixes[fix[1]].append(fix)

        for sheet_name, s_fixes in sorted(sheet_fixes.items()):
            if sheet_name not in wb.sheetnames:
                print(f"  ERROR: Sheet '{sheet_name}' not found in {filename}")
                continue

            ws = wb[sheet_name]
            print(f"\n  Sheet: {sheet_name} ({len(s_fixes)} fixes)")

            # Build ItemID -> row cache for this sheet
            id_to_row = {}
            for row in ws.iter_rows(min_row=2, max_col=1):
                if row[0].value:
                    id_to_row[str(row[0].value).strip()] = row[0].row

            for fix in s_fixes:
                item_id, sheet, col, old_text, new_text, severity = fix

                row = id_to_row.get(item_id)
                if not row:
                    print(f"    SKIP: ItemID {item_id} not found in sheet")
                    total_not_found += 1
                    results.append({"item_id": item_id, "status": "not_found"})
                    continue

                # Handle 'unknown' column - search all text columns
                if col == "unknown":
                    found = False
                    for try_col in ["QuestionText", "Hint1", "Hint2", "Hint3",
                                    "Option1", "Option2", "Option3", "Option4"]:
                        col_idx = COL_MAP[try_col]
                        cell_value = str(ws.cell(row=row, column=col_idx).value or "")
                        if safe_replace(cell_value, old_text, new_text) is not None:
                            col = try_col
                            found = True
                            break
                    if not found:
                        print(f"    SKIP: {item_id} - '{old_text}' not found in any column")
                        total_skipped += 1
                        results.append({"item_id": item_id, "status": "text_not_found",
                                       "old": old_text})
                        continue

                if dry_run:
                    # Just check if the text exists
                    col_idx = COL_MAP.get(col)
                    if col_idx:
                        cell_value = str(ws.cell(row=row, column=col_idx).value or "")
                        if safe_replace(cell_value, old_text, new_text) is not None:
                            print(f"    WOULD FIX: {item_id} [{col}] '{old_text}' -> '{new_text}'")
                            total_applied += 1
                        else:
                            print(f"    SKIP: {item_id} [{col}] '{old_text}' not found")
                            total_skipped += 1
                else:
                    success, detail = apply_fix(ws, row, col, old_text, new_text)
                    if success:
                        total_applied += 1
                        results.append({"item_id": item_id, "col": col,
                                       "old": old_text, "new": new_text, "status": "fixed"})
                    else:
                        total_skipped += 1
                        print(f"    SKIP: {item_id} [{col}] {detail}")
                        results.append({"item_id": item_id, "col": col,
                                       "old": old_text, "status": "skipped", "detail": detail})

        if not dry_run:
            wb.save(filepath)
            print(f"\n  Saved: {filepath}")
        wb.close()

    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"  Applied:   {total_applied}")
    print(f"  Skipped:   {total_skipped}")
    print(f"  Not found: {total_not_found}")
    print(f"  Total:     {total_applied + total_skipped + total_not_found}")

    if not dry_run and results:
        report_path = os.path.join(DATA_DIR, "fix_report.json")
        with open(report_path, "w") as f:
            json.dump({
                "applied": total_applied,
                "skipped": total_skipped,
                "not_found": total_not_found,
                "details": results
            }, f, indent=2)
        print(f"\n  Report saved: {report_path}")


if __name__ == "__main__":
    main()
