#!/usr/bin/env python3
"""
AI-powered quality review of questions.

Pass 1: Quality Check
- Spelling/grammar errors (Australian English)
- Verify correct answer is actually correct
- Check options make sense
- Flag formatting issues
- Flag missing data

Usage:
    source venv/bin/activate
    python -m scripts.bulk_import.review_questions --limit 10  # test with 10
    python -m scripts.bulk_import.review_questions --dry-run   # preview only
    python -m scripts.bulk_import.review_questions             # review all
    python -m scripts.bulk_import.review_questions --file "path/to/file.xlsx"
"""

import os
import sys
import argparse
import json
import time
from pathlib import Path
from datetime import datetime

import openpyxl
import anthropic

# Claude config
MODEL = "claude-haiku-4-5-20251001"
MAX_TOKENS = 500

# Rate limiting
DELAY_BETWEEN_CALLS = 0.3  # seconds

# Default files to review (Stage 1 master files)
DEFAULT_FILES = [
    "data/questions/clean/Krisite Stage One English Questions WORLD WISE.xlsx",
    "data/questions/clean/Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "data/questions/clean/Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
]

# Column mapping (21-column template format)
COL = {
    "ItemID": 0,
    "QuestionType": 1,
    "Subject": 2,
    "Category": 3,
    "Grade": 4,
    "Topic": 5,
    "QuestionText": 6,
    "Option1": 7,
    "Option2": 8,
    "Option3": 9,
    "Option4": 10,
    "Answer": 11,
    "Level": 12,
    "MediaType": 13,
    "ImageRequired": 14,
    "ImageDescription": 15,
    "Hint1": 16,
    "Hint2": 17,
    "Hint3": 18,
    "GetHelp": 19,
    "Notes": 20,
}


def get_api_key():
    """Get Anthropic API key from environment or .env file."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        env_path = Path(__file__).resolve().parent.parent.parent / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                if line.startswith("ANTHROPIC_API_KEY="):
                    api_key = line.split("=", 1)[1].strip()
                    break
    return api_key


def read_questions(filepath: str) -> list:
    """Read all questions from an Excel file. Returns list of dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    questions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        for row_idx in range(2, ws.max_row + 1):
            vals = [cell.value for cell in ws[row_idx]]

            # Skip empty rows
            item_id = vals[COL["ItemID"]] if len(vals) > COL["ItemID"] else None
            if not item_id:
                continue

            q_text = vals[COL["QuestionText"]] if len(vals) > COL["QuestionText"] else None
            if not q_text:
                continue

            question = {
                "file": Path(filepath).name,
                "sheet": sheet_name,
                "row": row_idx,
                "item_id": str(item_id),
                "question_type": str(vals[COL["QuestionType"]] or ""),
                "subject": str(vals[COL["Subject"]] or ""),
                "category": str(vals[COL["Category"]] or ""),
                "grade": str(vals[COL["Grade"]] or ""),
                "topic": str(vals[COL["Topic"]] or ""),
                "question_text": str(q_text),
                "option1": str(vals[COL["Option1"]] or "") if len(vals) > COL["Option1"] else "",
                "option2": str(vals[COL["Option2"]] or "") if len(vals) > COL["Option2"] else "",
                "option3": str(vals[COL["Option3"]] or "") if len(vals) > COL["Option3"] else "",
                "option4": str(vals[COL["Option4"]] or "") if len(vals) > COL["Option4"] else "",
                "answer": str(vals[COL["Answer"]] or "") if len(vals) > COL["Answer"] else "",
                "level": str(vals[COL["Level"]] or "") if len(vals) > COL["Level"] else "",
                "hint1": str(vals[COL["Hint1"]] or "") if len(vals) > COL["Hint1"] else "",
                "hint2": str(vals[COL["Hint2"]] or "") if len(vals) > COL["Hint2"] else "",
                "hint3": str(vals[COL["Hint3"]] or "") if len(vals) > COL["Hint3"] else "",
            }

            # Clean up "None" strings
            for k, v in question.items():
                if v == "None":
                    question[k] = ""

            questions.append(question)

    return questions


def review_question(client: anthropic.Anthropic, question: dict) -> dict:
    """Send a question to Claude for quality review. Returns issues found."""

    # Build options text
    options = []
    for i in range(1, 5):
        opt = question.get(f"option{i}", "")
        if opt:
            options.append(f"  Option {i}: {opt}")
    options_text = "\n".join(options) if options else "  (No options - True/False question)"

    # Build hints text
    hints = []
    for i in range(1, 4):
        hint = question.get(f"hint{i}", "")
        if hint:
            hints.append(f"  Hint {i}: {hint}")
    hints_text = "\n".join(hints) if hints else "  (No hints)"

    # Resolve the answer to show what it points to
    answer_raw = question['answer']
    answer_display = answer_raw
    if question['question_type'] in ('Select One', 'Select All'):
        try:
            ans_num = int(float(answer_raw))
            ans_text = question.get(f"option{ans_num}", "")
            answer_display = f"Option {ans_num} (which is: \"{ans_text}\")"
        except (ValueError, TypeError):
            pass

    prompt = f"""Review this educational question for primary school students (Australian English).

Question Type: {question['question_type']}
Subject: {question['subject']} - {question['category']}
Grade: Year {question['grade']}
Question: {question['question_text']}
Options:
{options_text}
Marked Correct Answer: {answer_display}
Hints:
{hints_text}

IMPORTANT: The "Marked Correct Answer" shows which option number is marked as correct AND the text of that option. Verify that this option is genuinely the correct answer to the question.

Check for these issues and respond in JSON format:
{{
  "answer_correct": true/false,
  "answer_issue": "explanation if wrong, empty string if correct",
  "spelling_errors": ["list of spelling/grammar errors in question text and options, empty array if none"],
  "question_clear": true/false,
  "question_issue": "explanation if unclear, empty string if clear",
  "options_issue": "any problems with the options, empty string if fine",
  "severity": "ok/minor/major",
  "notes": "any other observations, empty string if none"
}}

Rules:
- Use Australian English spelling (colour, favourite, behaviour, etc.)
- For True/False questions, the answer is "True" or "False"
- For Select All, the answer is like "1 and 3" meaning options 1 and 3 are both correct
- Be strict about spelling errors in the question text AND the options
- Only flag real issues, not style preferences
- Do NOT flag missing hints as an issue (hints are handled separately)
- "severity": "ok" means no issues, "minor" means typos/small formatting issues, "major" means wrong answer or question doesn't make sense

Respond with ONLY valid JSON, no other text."""

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=[
                {"role": "user", "content": prompt}
            ],
        )

        result_text = response.content[0].text.strip()
        # Clean up potential markdown formatting
        if result_text.startswith("```"):
            result_text = result_text.split("\n", 1)[1]
            if result_text.endswith("```"):
                result_text = result_text[:-3]
            result_text = result_text.strip()

        result = json.loads(result_text)
        return result

    except json.JSONDecodeError as e:
        return {"error": f"Failed to parse AI response: {e}", "severity": "error"}
    except anthropic.APIStatusError as e:
        if e.status_code == 529:  # Overloaded
            return {"error": "API overloaded", "severity": "retry"}
        return {"error": str(e), "severity": "error"}
    except Exception as e:
        return {"error": str(e), "severity": "error"}


def main():
    parser = argparse.ArgumentParser(
        description="AI-powered question quality review"
    )
    parser.add_argument(
        "--file", "-f",
        type=str,
        help="Specific file to review (default: all Stage 1 master files)"
    )
    parser.add_argument(
        "--limit", "-l",
        type=int,
        default=0,
        help="Limit number of questions to review (for testing)"
    )
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Preview what would be reviewed without making API calls"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output file for report (default: data/questions/review_report_YYYYMMDD.json)"
    )
    parser.add_argument(
        "--resume", "-r",
        type=str,
        default=None,
        help="Resume from a previous report file (skips already-reviewed questions)"
    )
    parser.add_argument(
        "--offset",
        type=int,
        default=0,
        help="Skip first N questions (for resuming from a known position)"
    )

    args = parser.parse_args()

    # Get API key
    api_key = get_api_key()
    if not api_key and not args.dry_run:
        print("Error: ANTHROPIC_API_KEY not found")
        sys.exit(1)

    # Determine files to review
    if args.file:
        files = [args.file]
    else:
        files = DEFAULT_FILES

    print("=" * 60)
    print("QUESTION QUALITY REVIEW - Pass 1")
    print("=" * 60)

    # Read all questions
    all_questions = []
    for f in files:
        if not Path(f).exists():
            print(f"Warning: {f} not found, skipping")
            continue
        print(f"\nReading {Path(f).name}...")
        questions = read_questions(f)
        print(f"  Found {len(questions)} questions")
        all_questions.extend(questions)

    print(f"\nTotal questions to review: {len(all_questions)}")

    if args.offset > 0:
        all_questions = all_questions[args.offset:]
        print(f"Skipping first {args.offset} (--offset)")

    if args.limit > 0:
        all_questions = all_questions[:args.limit]
        print(f"Limiting to first {args.limit}")

    if args.dry_run:
        print("\n[DRY RUN - No API calls]")
        print(f"\nWould review {len(all_questions)} questions")
        print(f"Estimated cost: ${len(all_questions) * 0.003:.2f}")
        print(f"\nSample questions:")
        for q in all_questions[:5]:
            print(f"  {q['item_id']}: {q['question_text'][:80]}")
        return

    # Initialize Claude client
    client = anthropic.Anthropic(api_key=api_key)

    # Resume from previous report if specified
    reviewed_ids = set()
    results = {
        "reviewed_at": datetime.now().isoformat(),
        "total_questions": len(all_questions),
        "issues": [],
        "stats": {
            "ok": 0,
            "minor": 0,
            "major": 0,
            "error": 0,
        }
    }

    if args.resume and Path(args.resume).exists():
        with open(args.resume) as f:
            prev_report = json.load(f)
        # Load previous stats and issues
        results["stats"] = prev_report["stats"]
        results["issues"] = prev_report["issues"]
        # Track which ItemIDs were already reviewed
        for issue in prev_report["issues"]:
            reviewed_ids.add(issue["item_id"])
        # Also count OKs (they aren't in issues list)
        reviewed_ids_from_stats = results["stats"]["ok"] + results["stats"]["minor"] + results["stats"]["major"] + results["stats"]["error"]
        print(f"\nResuming from previous report: {args.resume}")
        print(f"  Already reviewed: {reviewed_ids_from_stats}")
        # Filter out already-reviewed questions
        all_questions = [q for q in all_questions if q["item_id"] not in reviewed_ids]
        # For OK questions not in issues, skip by offset
        skip_count = reviewed_ids_from_stats - len(reviewed_ids)
        if skip_count > 0:
            all_questions = all_questions[skip_count:]
        print(f"  Remaining: {len(all_questions)}")

    # Output file
    output_path = args.output or args.resume or f"data/questions/review_report_{datetime.now().strftime('%Y%m%d_%H%M')}.json"

    # Review questions
    print(f"\nReviewing {len(all_questions)} questions...")
    print("(Resumable - saves every 50 questions)")
    print("-" * 60)

    consecutive_errors = 0

    for i, q in enumerate(all_questions, 1):
        print(f"[{i}/{len(all_questions)}] {q['item_id']}: {q['question_text'][:60]}...", end=" ", flush=True)

        review = review_question(client, q)
        severity = review.get("severity", "error")

        if severity == "ok":
            print("OK")
            results["stats"]["ok"] += 1
            consecutive_errors = 0
        elif severity == "minor":
            print(f"MINOR")
            issues_found = []
            if review.get("spelling_errors"):
                issues_found.append(f"Spelling: {', '.join(review['spelling_errors'])}")
            if review.get("options_issue"):
                issues_found.append(f"Options: {review['options_issue']}")
            if review.get("notes") and review["notes"]:
                issues_found.append(f"Note: {review['notes']}")
            for issue in issues_found:
                print(f"    {issue}")
            results["stats"]["minor"] += 1
            results["issues"].append({**q, "review": review})
            consecutive_errors = 0
        elif severity == "major":
            print(f"MAJOR")
            if review.get("answer_issue"):
                print(f"    ANSWER: {review['answer_issue']}")
            if review.get("question_issue"):
                print(f"    QUESTION: {review['question_issue']}")
            if review.get("spelling_errors"):
                print(f"    SPELLING: {', '.join(review['spelling_errors'])}")
            results["stats"]["major"] += 1
            results["issues"].append({**q, "review": review})
            consecutive_errors = 0
        elif severity == "retry":
            # API overloaded - wait and retry
            print(f"OVERLOADED - waiting 30s...", flush=True)
            time.sleep(30)
            review = review_question(client, q)
            severity = review.get("severity", "error")
            if severity == "ok":
                print(f"    Retry OK")
                results["stats"]["ok"] += 1
            elif severity in ("minor", "major"):
                print(f"    Retry {severity.upper()}")
                results["stats"][severity] += 1
                results["issues"].append({**q, "review": review})
            else:
                print(f"    Retry failed - skipping")
                results["stats"]["error"] += 1
                consecutive_errors += 1
        else:
            print(f"ERROR: {review.get('error', 'unknown')}")
            results["stats"]["error"] += 1
            consecutive_errors += 1

        # Save incrementally every 50 questions
        if i % 50 == 0:
            with open(output_path, "w") as f:
                json.dump(results, f, indent=2)
            print(f"\n--- Progress: {i}/{len(all_questions)} | OK: {results['stats']['ok']} | Minor: {results['stats']['minor']} | Major: {results['stats']['major']} ---\n")

        # Stop on too many consecutive errors (but not overload retries)
        if consecutive_errors >= 20:
            print(f"\n20 consecutive errors - stopping.")
            break

        # Rate limiting
        time.sleep(DELAY_BETWEEN_CALLS)

    # Save final report
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2)

    # Summary
    print("\n" + "=" * 60)
    print("REVIEW COMPLETE")
    print("=" * 60)
    print(f"\nOK: {results['stats']['ok']}")
    print(f"Minor issues: {results['stats']['minor']}")
    print(f"Major issues: {results['stats']['major']}")
    print(f"Errors: {results['stats']['error']}")
    print(f"\nReport saved to: {output_path}")

    if results["stats"]["major"] > 0:
        print(f"\n⚠ {results['stats']['major']} MAJOR issues found - review the report!")


if __name__ == "__main__":
    main()
