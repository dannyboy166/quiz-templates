"""
Generate progressive hints for questions that don't have any.

Uses Claude API (Haiku 4.5) with subject-specific prompts and
curated examples from Kristie's existing hints.

Usage:
    source venv/bin/activate

    # Phase 1: Test 50 questions, review output
    python -m scripts.bulk_import.generate_hints --limit 50

    # Dry run (just show which questions need hints)
    python -m scripts.bulk_import.generate_hints --dry-run

    # Process specific sheet
    python -m scripts.bulk_import.generate_hints --sheet "Grammar" --limit 20

    # Full run
    python -m scripts.bulk_import.generate_hints

    # Apply generated hints back to spreadsheets
    python -m scripts.bulk_import.generate_hints --apply
"""

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime

import anthropic
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# === Configuration ===

DATA_DIR = "data/questions"
CLEAN_DIR = os.path.join(DATA_DIR, "clean")
OUTPUT_DIR = os.path.join(DATA_DIR, "hints")

MODEL = "claude-haiku-4-5-20251001"

SHEET_TO_FILE = {
    "Phonological Awareness": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Phonics": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Spelling": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Punctuation": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Grammar": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Vocabulary": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Reading Comprehension": "Krisite Stage One English Questions WORLD WISE.xlsx",
    "Number and Algebra": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "Measurement & Space": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "Statistics & Probability": "Kristie Stage One Mathematics Questions WORLD WISE.xlsx",
    "HSIE": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "Science & technology": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "Creative Arts": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
    "PD H PE": "Kristie Stage One Other Subjects Questions WORLD WISE.xlsx",
}

# Subject groups for different system prompts
SUBJECT_GROUPS = {
    "english": ["Spelling", "Grammar", "Punctuation", "Phonics", "Vocabulary",
                "Phonological Awareness"],
    "reading": ["Reading Comprehension"],
    "maths": ["Number and Algebra", "Measurement & Space", "Statistics & Probability"],
    "other": ["HSIE", "Science & technology", "Creative Arts", "PD H PE"],
}

# === System Prompts ===

BASE_RULES = """Rules:
- Write in simple, child-friendly Australian English (colour, favourite, recognise)
- Keep each hint to 1-2 short sentences (under 120 characters preferred, max 150)
- Use a friendly, encouraging tone
- NEVER say the answer directly in any hint
- NEVER reference images, arrows, circles, colours, highlights or visual elements — hints are TEXT ONLY
- NEVER say "look at the image" or "click on" or "have the word X displayed"
- For True/False questions: guide thinking without saying "true" or "false"
- For Select All questions: Hint 3 should remind them to select more than one answer
- Hint 1 should work even if the student hasn't read the options yet

Respond with ONLY a JSON object (no markdown, no explanation):
{"hint1": "...", "hint2": "...", "hint3": "..."}"""

ENGLISH_SYSTEM = f"""You are generating progressive hints for Australian primary school students (Years 1-2, ages 5-8) for English literacy questions.

Generate exactly 3 hints that progressively help the student:
- Hint 1: A gentle nudge — define the concept, teach the rule, or give a general direction.
- Hint 2: A stronger clue — narrow it down, apply the rule to this specific question, or eliminate wrong answers.
- Hint 3: Almost gives it away — make the correct answer very obvious through a leading question or near-complete scaffold.

Examples of good hints:

Question: "Which sentence has correct punctuation?" Options: He likes to run / He likes to run. / he likes to run.
Hint 1: "remember that a sentence needs a capital letter at the start and a full stop at the end."
Hint 2: "check each answer — does it start with a capital letter AND end with a full stop?"
Hint 3: "only one answer has both a capital letter at the start and a full stop at the end."

Question: "What is the contraction for cannot?" Options: can not / can't / cant
Hint 1: "a contraction joins two words together using an apostrophe (') to show missing letters."
Hint 2: "the two words are 'can' and 'not'. When joined, some letters are replaced by an apostrophe."
Hint 3: "which answer has an apostrophe in the right place?"

Question: "Which word belongs to the 'at' word family?" Options: sun / hat / dog / map
Hint 1: "words in the same word family share the same ending sound and letters."
Hint 2: "the 'at' family means the word ends in the letters a-t."
Hint 3: "say each word slowly. Which one ends with the 'at' sound?"

Question: "Which two words rhyme?" Options: dog / pig / mat / jet (Answer: 1 and 2)
Hint 1: "words that rhyme have the same ending sound."
Hint 2: "say each word out loud. Listen for two words that end with the same sound."
Hint 3: "listen to the ending sound of 'dog' — which other word ends the same way?"

Question: "Choose the missing word: The cat ___ on the mat." Options: sat / in / and (Answer: 1)
Hint 1: "read the sentence and think about which word makes it sound right."
Hint 2: "the missing word tells us what the cat did."
Hint 3: "what is something a cat does on a mat? It starts with 's'."

{BASE_RULES}"""

READING_SYSTEM = f"""You are generating progressive hints for Australian primary school students (Years 1-2, ages 5-8) for reading comprehension questions.

These questions are about a text passage the student has already read. The hints should guide them back to the text.

Generate exactly 3 hints:
- Hint 1: A general comprehension strategy — look back at the text, find key words, etc.
- Hint 2: Point them to where in the text they can find the answer (without saying the answer).
- Hint 3: Almost gives it away — rephrase the answer-containing part of the text as a question.

Examples of good hints:

Question: "Where did Ben go with Mum?" Options: beach / pool / swim / shop (Answer: beach)
Hint 1: "the answer is in the text. Read through it again carefully."
Hint 2: "look for the part that tells you where Ben and Mum went together."
Hint 3: "read the sentence that mentions Mum. Where did they go?"

Question: "Why did the puppy jump on Pam's lap?" Options: It was scared / It was happy / It was hungry
Hint 1: "this answer is not written directly in the text. You need to think about the clues."
Hint 2: "think about how the puppy was feeling when it saw Pam."
Hint 3: "the puppy wagged its tail when it saw Pam. What does that tell you about how it felt?"

{BASE_RULES}"""

MATHS_SYSTEM = f"""You are generating progressive hints for Australian primary school students (Years 1-2, ages 5-8) for mathematics questions.

Generate exactly 3 hints that progressively help:
- Hint 1: Explain the concept or strategy needed (place value, counting, addition, etc.)
- Hint 2: Walk them through the first step or give a concrete strategy to try.
- Hint 3: Almost gives it away — show most of the working or give a very strong clue.

Examples of good hints:

Question: "Which number is represented as 2 tens and 4 ones?" Options: 40 / 24 / 42 / 20 (Answer: 24)
Hint 1: "1 ten is the same as 10 ones. Think about how many ones are in 2 tens."
Hint 2: "2 tens is 20. Now add the 4 extra ones."
Hint 3: "20 + 4 = ?"

Question: "3 groups of 5 make ___" Options: 5 / 10 / 15 / 20 (Answer: 15)
Hint 1: "groups of means you need to add the same number that many times."
Hint 2: "3 groups of 5 means 5 + 5 + 5."
Hint 3: "5 + 5 = 10. Now add one more 5."

Question: "A shape has 4 sides and 4 corners. Which shape could it be?" Options: circle / square / triangle / hexagon
Hint 1: "count the sides and corners of each shape you know."
Hint 2: "a circle has no straight sides. A triangle has 3 sides. Which shape has exactly 4?"
Hint 3: "think of a shape that looks like a box or a window. It has 4 equal sides."

Question: "Order these numbers from largest to smallest: 34, 78, 12, 56"
Hint 1: "look at the tens digit first — the bigger the tens digit, the bigger the number."
Hint 2: "which number has the biggest tens digit? That goes first."
Hint 3: "7 tens is the biggest, then 5 tens, then 3 tens, then 1 ten."

{BASE_RULES}"""

OTHER_SYSTEM = f"""You are generating progressive hints for Australian primary school students (Years 1-2, ages 5-8) for general knowledge questions (science, history, geography, health, creative arts).

Generate exactly 3 hints:
- Hint 1: Define the key concept or give background knowledge needed to answer.
- Hint 2: Narrow it down — eliminate obviously wrong answers or give a specific fact.
- Hint 3: Almost gives it away — ask a leading question that points directly to the answer.

Examples of good hints:

Question: "Rocks are living things. True or false?" (Answer: False)
Hint 1: "a living thing is something that is alive. Living things grow, move and need food and water."
Hint 2: "think about whether a rock grows, moves or needs food."
Hint 3: "does a rock grow and change? Does it need food or water?"

Question: "Which one is something from the past?" Options: Phone / Laptop / Watch / Horse and cart (Answer: 4)
Hint 1: "the past means a long time ago, before modern technology was invented."
Hint 2: "phones, laptops and watches are things we still use today."
Hint 3: "which one was used for transport a long time ago, before cars were invented?"

Question: "Before crossing a road, what should you do?" Options: Run across / Skip across / Stop, look, listen and think
Hint 1: "road safety means being careful around cars and traffic."
Hint 2: "we should never rush across a road without checking first."
Hint 3: "what should you do with your eyes and ears before you step onto the road?"

Question: "A line can be straight, wiggly or curved. True or false?" (Answer: True)
Hint 1: "think about all the different types of lines you can draw."
Hint 2: "can you draw a line that is not straight? What would it look like?"
Hint 3: "artists use many different types of lines — straight ones, wavy ones and curved ones."

{BASE_RULES}"""

SYSTEM_PROMPTS = {
    "english": ENGLISH_SYSTEM,
    "reading": READING_SYSTEM,
    "maths": MATHS_SYSTEM,
    "other": OTHER_SYSTEM,
}


def get_subject_group(sheet_name):
    """Get the prompt group for a sheet."""
    for group, sheets in SUBJECT_GROUPS.items():
        if sheet_name in sheets:
            return group
    return "other"


def read_questions_needing_hints(sheet_filter=None):
    """Read all questions that don't have Hint1 populated."""
    questions = []

    for sheet_name, filename in SHEET_TO_FILE.items():
        if sheet_filter and sheet_name != sheet_filter:
            continue

        filepath = os.path.join(CLEAN_DIR, filename)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2):
            if not row[0].value:
                continue

            # Check if Hint1 is empty
            h1 = row[16].value if len(row) > 16 else None
            if h1 and str(h1).strip():
                continue  # Already has hints

            item_id = str(row[0].value).strip()
            q_type = str(row[1].value or "").strip()
            q_text = str(row[6].value or "").strip()
            options = []
            for i in range(7, 11):
                opt = str(row[i].value or "").strip() if len(row) > i else ""
                if opt:
                    options.append(opt)
            answer = str(row[11].value or "").strip() if len(row) > 11 else ""
            subject = str(row[2].value or "").strip()
            category = str(row[3].value or "").strip()
            topic = str(row[5].value or "").strip()

            if not q_text:
                continue

            questions.append({
                "item_id": item_id,
                "sheet": sheet_name,
                "subject_group": get_subject_group(sheet_name),
                "question_type": q_type,
                "question": q_text,
                "options": options,
                "answer": answer,
                "subject": subject,
                "category": category,
                "topic": topic,
            })

        wb.close()

    return questions


def build_user_prompt(q):
    """Build the user message for a single question."""
    parts = [f'Question type: {q["question_type"]}']
    parts.append(f'Subject: {q["subject"]} > {q["category"]}')
    parts.append(f'Question: "{q["question"]}"')

    if q["options"]:
        opts_str = " / ".join(f'{i+1}. {opt}' for i, opt in enumerate(q["options"]))
        parts.append(f'Options: {opts_str}')

    if q["answer"]:
        parts.append(f'Correct answer: {q["answer"]}')

    return "\n".join(parts)


def generate_hint(client, question, system_prompt):
    """Call Claude API to generate hints for one question."""
    user_msg = build_user_prompt(question)

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=300,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )

        text = response.content[0].text.strip()

        # Parse JSON from response
        # Handle potential markdown wrapping
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\n?", "", text)
            text = re.sub(r"\n?```$", "", text)

        result = json.loads(text)

        return {
            "hint1": result.get("hint1", ""),
            "hint2": result.get("hint2", ""),
            "hint3": result.get("hint3", ""),
            "tokens_in": response.usage.input_tokens,
            "tokens_out": response.usage.output_tokens,
        }

    except json.JSONDecodeError as e:
        return {"error": f"JSON parse error: {e}", "raw": text}
    except Exception as e:
        return {"error": str(e)}


def validate_hints(hints, question):
    """Check generated hints for quality issues."""
    issues = []

    for key in ["hint1", "hint2", "hint3"]:
        h = hints.get(key, "")
        if not h:
            issues.append(f"{key} is empty")
            continue
        if len(h) < 10:
            issues.append(f"{key} too short ({len(h)} chars)")
        if len(h) > 200:
            issues.append(f"{key} too long ({len(h)} chars)")

    # Check answer leak
    answer = question.get("answer", "")
    if answer and len(answer) > 1:
        for key in ["hint1", "hint2", "hint3"]:
            h = hints.get(key, "").lower()
            ans_lower = answer.lower().strip()
            # For option-based answers (1, 2, 1 and 3, True, False) skip - too short/common
            if ans_lower in ("1", "2", "3", "4", "true", "false",
                            "1 and 2", "1 and 3", "1 and 4", "2 and 3", "2 and 4", "3 and 4"):
                continue
            # Check if the full answer text appears in the hint
            if ans_lower in h:
                issues.append(f"{key} may leak the answer (contains '{answer}')")
        # Also check if hint says "the answer is" or "that's your answer"
        for key in ["hint1", "hint2", "hint3"]:
            h = hints.get(key, "").lower()
            if "the answer is" in h or "your answer" in h:
                issues.append(f"{key} may directly state the answer")

    # Check for visual references (UI/image instructions, not educational use of the words)
    visual_phrases = [
        "look at the image", "look at the picture", "look at the diagram",
        "click on", "click the", "tap on", "tap the",
        "the highlighted", "the underlined", "the circled word",
        "draw a line", "draw an arrow", "colour in",
        "in the picture", "in the image", "in the diagram",
    ]
    for key in ["hint1", "hint2", "hint3"]:
        h = hints.get(key, "").lower()
        for vp in visual_phrases:
            if vp in h:
                issues.append(f"{key} references visual element: '{vp}'")
                break

    return issues


def save_results(results, output_path):
    """Save results incrementally."""
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2)


def apply_hints_to_spreadsheets(output_path):
    """Write generated hints back to the Excel files."""
    with open(output_path) as f:
        results = json.load(f)

    # Group by file
    by_file = defaultdict(list)
    for r in results:
        if "error" in r:
            continue
        sheet = r["sheet"]
        filename = SHEET_TO_FILE.get(sheet)
        if filename:
            by_file[filename].append(r)

    for filename, items in by_file.items():
        filepath = os.path.join(CLEAN_DIR, filename)
        print(f"  Opening: {filename} ({len(items)} hints to write)")

        wb = openpyxl.load_workbook(filepath)

        # Group by sheet
        by_sheet = defaultdict(list)
        for item in items:
            by_sheet[item["sheet"]].append(item)

        for sheet_name, sheet_items in by_sheet.items():
            ws = wb[sheet_name]

            # Build row lookup
            id_to_row = {}
            for row in ws.iter_rows(min_row=2, max_col=1):
                if row[0].value:
                    id_to_row[str(row[0].value).strip()] = row[0].row

            applied = 0
            for item in sheet_items:
                row_num = id_to_row.get(item["item_id"])
                if not row_num:
                    continue

                ws.cell(row=row_num, column=17).value = item.get("hint1", "")
                ws.cell(row=row_num, column=18).value = item.get("hint2", "")
                ws.cell(row=row_num, column=19).value = item.get("hint3", "")
                applied += 1

            print(f"    {sheet_name}: {applied} hints written")

        wb.save(filepath)
        wb.close()
        print(f"  Saved: {filepath}")


def main():
    parser = argparse.ArgumentParser(description="Generate hints for questions")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, no API calls")
    parser.add_argument("--limit", type=int, help="Limit number of questions to process")
    parser.add_argument("--sheet", type=str, help="Process only this sheet")
    parser.add_argument("--apply", action="store_true", help="Apply hints from output file to spreadsheets")
    parser.add_argument("--output", type=str, help="Output file path")
    parser.add_argument("--resume", action="store_true", help="Skip questions already in output file")
    parser.add_argument("--model", type=str, default=MODEL, help=f"Model to use (default: {MODEL})")
    args = parser.parse_args()

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Default output path
    if not args.output:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        args.output = os.path.join(OUTPUT_DIR, f"hints_{timestamp}.json")

    # Apply mode
    if args.apply:
        # Find most recent output file
        if not os.path.exists(args.output):
            # Try to find latest
            files = sorted(
                [f for f in os.listdir(OUTPUT_DIR) if f.startswith("hints_") and f.endswith(".json")],
                reverse=True,
            )
            if files:
                args.output = os.path.join(OUTPUT_DIR, files[0])
            else:
                print("ERROR: No hints output file found")
                sys.exit(1)

        print(f"Applying hints from: {args.output}")
        apply_hints_to_spreadsheets(args.output)
        return

    # Read questions
    print("Reading questions needing hints...")
    questions = read_questions_needing_hints(sheet_filter=args.sheet)
    print(f"  Found: {len(questions)} questions without hints")

    # Group by subject
    by_group = defaultdict(list)
    for q in questions:
        by_group[q["subject_group"]].append(q)

    print("  By group:")
    for group, items in sorted(by_group.items()):
        print(f"    {group}: {len(items)}")

    if args.limit:
        import random
        random.seed(42)

        # Take evenly from each SHEET (not group) for good coverage
        sampled = []
        by_sheet = defaultdict(list)
        for q in questions:
            by_sheet[q["sheet"]].append(q)

        per_sheet = max(1, args.limit // len(by_sheet))
        for sheet_name in sorted(by_sheet.keys()):
            sheet_qs = by_sheet[sheet_name]
            # Random sample from each sheet, picking different question types
            sample_size = min(per_sheet, len(sheet_qs))
            sampled.extend(random.sample(sheet_qs, sample_size))

        # If we still need more to reach the limit, fill from largest sheets
        if len(sampled) < args.limit:
            remaining_pool = [q for q in questions if q not in sampled]
            extra = min(args.limit - len(sampled), len(remaining_pool))
            sampled.extend(random.sample(remaining_pool, extra))

        questions = sampled[:args.limit]
        print(f"  Sampled: {len(questions)} questions (--limit {args.limit})")

    if args.dry_run:
        print(f"\nDRY RUN — would generate hints for {len(questions)} questions")
        print(f"\nSample questions:")
        for q in questions[:10]:
            print(f"  [{q['sheet']}] {q['item_id']}: {q['question'][:60]}")
        return

    # Resume support
    existing_ids = set()
    results = []
    if args.resume and os.path.exists(args.output):
        with open(args.output) as f:
            results = json.load(f)
        existing_ids = {r["item_id"] for r in results}
        print(f"  Resuming: {len(existing_ids)} already done")
        questions = [q for q in questions if q["item_id"] not in existing_ids]
        print(f"  Remaining: {len(questions)}")

    if not questions:
        print("Nothing to process!")
        return

    # Estimate cost
    est_tokens_in = len(questions) * 800  # ~800 input tokens per question
    est_tokens_out = len(questions) * 100  # ~100 output tokens per question
    # Haiku: $0.80/MTok in, $4/MTok out
    est_cost = (est_tokens_in * 0.80 + est_tokens_out * 4.0) / 1_000_000
    print(f"\n  Estimated cost: ${est_cost:.2f} ({len(questions)} questions)")
    print(f"  Model: {args.model}")
    print(f"  Output: {args.output}")
    print()

    # Generate
    client = anthropic.Anthropic()
    total_tokens_in = 0
    total_tokens_out = 0
    errors = 0
    quality_issues = 0

    for i, q in enumerate(questions):
        system_prompt = SYSTEM_PROMPTS[q["subject_group"]]

        result = generate_hint(client, q, system_prompt)

        # Add metadata
        result["item_id"] = q["item_id"]
        result["sheet"] = q["sheet"]
        result["question"] = q["question"]
        result["question_type"] = q["question_type"]

        if "error" in result:
            errors += 1
            print(f"  ERROR [{q['item_id']}]: {result['error']}")
        else:
            # Validate
            issues = validate_hints(result, q)
            if issues:
                result["quality_issues"] = issues
                quality_issues += 1

            total_tokens_in += result.get("tokens_in", 0)
            total_tokens_out += result.get("tokens_out", 0)

        results.append(result)

        # Progress
        if (i + 1) % 10 == 0 or i == len(questions) - 1:
            cost = (total_tokens_in * 0.80 + total_tokens_out * 4.0) / 1_000_000
            print(f"  [{i+1}/{len(questions)}] Cost so far: ${cost:.3f} | Errors: {errors} | Quality issues: {quality_issues}")

            # Save incrementally
            save_results(results, args.output)

        # Rate limiting: 50 requests per minute for Haiku
        if (i + 1) % 45 == 0:
            time.sleep(5)

    # Final save
    save_results(results, args.output)

    # Summary
    cost = (total_tokens_in * 0.80 + total_tokens_out * 4.0) / 1_000_000
    print(f"\n{'='*60}")
    print(f"COMPLETE")
    print(f"{'='*60}")
    print(f"  Questions processed: {len(questions)}")
    print(f"  Successful: {len(questions) - errors}")
    print(f"  Errors: {errors}")
    print(f"  Quality issues: {quality_issues}")
    print(f"  Total cost: ${cost:.3f}")
    print(f"  Output: {args.output}")

    # Show sample
    print(f"\n  Sample outputs:")
    for r in results[:5]:
        if "error" not in r:
            print(f"\n    [{r['sheet']}] {r['question'][:50]}")
            print(f"      H1: {r.get('hint1', '')}")
            print(f"      H2: {r.get('hint2', '')}")
            print(f"      H3: {r.get('hint3', '')}")
            if r.get("quality_issues"):
                print(f"      ⚠️  {r['quality_issues']}")


if __name__ == "__main__":
    main()
